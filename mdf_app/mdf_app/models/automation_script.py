from django.http import Http404
from django.shortcuts import get_object_or_404
from openpyxl import load_workbook
from . import Module, Coursework, Exam, Project, Staff, Duty, AdministrativeDuty
from .choices import Semester, ElectiveOptions

filename = "mdf_app/models/MDFS_20 06 22-1.xlsx"

def batch_upload(filename: str):
    wb = load_workbook(filename=filename)
    sheet = wb.get_sheet_by_name(name="MDFs")
    for line in sheet.iter_rows(2, values_only=True):
        if not line[0]:
            continue
        print(str(line[0] or ''))
        try:
            m = get_object_or_404(Module, code = str(line[0] or '').strip())
            print("found " + str(line[0] or ''))
        except Http404 as e:
            print("failed to find " + str(line[0] or ''))
            m = Module()
            m.code = str(line[0] or '')
        m.name = str(line[1] or '')
        m.credits = str(line[2] or 0)
        m.semester = Semester.values[Semester.labels.index(str(line[3] or '1/2'))]
        if line[4] == 'NE':
            line[4] = 'Compulsory'
        if line[4] == 'PE':
            line[4] = 'Optional'
        if line[4] == 'EO':
            line[4] = 'Elective'
        m.elective = ElectiveOptions.values[ElectiveOptions.labels.index(str(line[4] or 'Compulsory'))]
        m.level = str(line[5] or '1')
        m.aims = str(line[6] or '')
        m.learning_outcomes = str(line[7] or '')
        m.syllabus = str(line[8] or '')
        m.criteria = str(line[9] or '')
        m.feedback = str(line[10] or '')
        m.comments = str(line[11] or '')
        m.reading = str(line[12] or '')
        m.lecture_hours = str(line[14] or 0)
        m.tutorial_hours = str(line[15] or 0)
        m.assignment_hours = str(line[16] or 0)
        m.lab_hours = str(line[17] or 0)
        m.study_hours = str(line[18] or 0)
        m.save()
        for req in str(line[13] or '').split(','):
            if req == '':
                continue
            #print('adding reqs')
            code = req.strip()
            try:
                m.required_modules.add(get_object_or_404(Module, code=code))
            except Http404 as e:
                print("failed to find " + code)
                newm = Module()
                newm.code = code
                newm.save()
                m.required_modules.add(newm)
        m.save()
    wb.close()


def batch_upload_cw(filename: str):
    wb = load_workbook(filename=filename)
    sheet = wb.get_sheet_by_name(name="Coursework")
    Coursework.objects.all().delete()
    Exam.objects.all().delete()
    Project.objects.all().delete()
    for line in sheet.iter_rows(2, values_only=True):
        if not line[0]:
            continue

        if line[1] == 'C':
            c = Coursework()
            c.module = get_object_or_404(Module, code=line[0])
            c.semester= str(line[2] or '')
            c.week=str(line[3] or '')
            c.weight=float(line[4] or 0.0)
            c.information=str(line[5] or '')
            c.save()
        elif line[1] == 'E':
            e = Exam()
            e.module = get_object_or_404(Module, code=line[0])
            e.semester = str(line[2] or '')
            e.week = str(line[3] or '')
            e.weight = float(line[4] or 0.0)
            e.information = str(line[5] or '')
            e.duration = float(line[6] or 0.0)
            e.save()
        elif line[1] == 'P':
            p = Project()
            p.module = get_object_or_404(Module, code=line[0])
            p.semester= str(line[2] or '')
            p.week=str(line[3] or '')
            p.weight=float(line[4] or 0.0)
            p.information=str(line[5] or '')
            p.save()
    wb.close()
    
def batch_upload_staff_from_excel(filename: str):
    wb = load_workbook(filename=filename)
    for sheet in wb.worksheets:
            for line in sheet.iter_rows(2, values_only=True):
                if not line[0]:
                    continue
                try:
                    s = get_object_or_404(Staff, id_number = str(line[0] or ''))
                except Http404 as e:
                    s = Staff()
                    s.id_number = str(line[0] or '')
                s.title = str(line[1] or '')
                s.forename = str(line[2] or '')
                s.surname = str(line[3] or '')
                s.save()
               
    wb.close()
   
def update_all_staff_from_db():
    import pyodbc 
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=ENG-AENEA;'
                          'Database=FOE;'
                          'UID=mdf-app;'
                          'PWD=kuwuca-34;'
                          )
               
    cursor = conn.cursor()

    items = cursor.execute("SELECT [cntPersonnelId]"
                   ",[strTitle]"
                   ",[strSalutation]"
                   ",[strForename]"
                   ",[strSurname]"
                   ",[strJobTitle]"
                   ",[strEmailAddress]"
                   ",[lngPersonID]"
                   "FROM [FOE].[MDF].[vtblPersonnel]").fetchall()
                   
    for i in items:
        try:
                s = get_object_or_404(Staff, personnel_id = i[0])
                print("Got " + s.full_name())
        except Http404 as e:
            s = Staff()
            s.personnel_id = i[0]
            print("Creating %s %s %s" % (i[1], i[2], i[4]))
        s.title = i[1]
        s.salutation = i[2]
        s.forename= i[3]
        s.surname = i[4]
        s.job_title = i[5]
        s.email = i[6]
        s.save()
        
def dump_IC_to_sheet(filename):
    with open(filename, "w") as f:
        f.write("Code,Class title,Registrar,Email\n")
        duties = Duty.objects.all()
        for d in duties:
            if d.role == '1':
                f.write("%s,%s,%s,%s\n" % (d.module.code, d.module.name, d.assignee.full_name(), d.assignee.email))
                
                

