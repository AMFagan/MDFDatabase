from django.http import Http404
from django.shortcuts import get_object_or_404
from openpyxl import load_workbook
from . import Module, Coursework, Exam, Project
from .choices import Semester, ElectiveOptions


def batch_upload(filename: str):
    wb = load_workbook(filename=filename)
    for sheet in wb.worksheets:
            for line in sheet.iter_rows(2, values_only=True):
                if not line[0]:
                    continue
                try:
                    m = get_object_or_404(Module, code = str(line[0] or ''))
                except Http404 as e:
                    m = Module()
                    m.code = str(line[0] or '')
                m.name = str(line[1] or '')
                m.credits = str(line[2] or 0)
                m.semester = Semester.values[Semester.labels.index(str(line[3] or '1/2'))]
                m.elective = ElectiveOptions.values[ElectiveOptions.labels.index(str(line[4] or 'NE'))]
                m.level = str(line[5] or '1')
                m.aims = str(line[6] or '')
                m.learning_outcomes = str(line[7] or '')
                m.syllabus = str(line[8] or '')
                m.criteria = str(line[9] or '')
                m.feedback = str(line[10] or '')
                m.comments = str(line[11] or '')
                m.reading = str(line[12] or '')
                m.lecture_hours = str(line[14] or 0)
                m.tutorial_hours = str(line[15] or 0)
                m.assignment_hours = str(line[16] or 0)
                m.lab_hours = str(line[17] or 0)
                m.study_hours = str(line[18] or 0)
                m.save()
                for req in str(line[13] or '').split(','):
                    if req == '':
                        continue
                    #print('adding reqs')
                    code = req.strip()
                    try:
                        m.required_modules.add(get_object_or_404(Module, code=code))
                    except Http404 as e:
                        newm = Module()
                        newm.code = code
                        newm.save()
                        m.required_modules.add(newm)
                m.save()
    wb.close()


def batch_upload_cw(filename: str):
    wb = load_workbook(filename=filename)
    for sheet in wb.worksheets:
            for line in sheet.iter_rows(2, values_only=True):
                if not line[0]:
                    continue

                if line[1] == 'C':
                    c = Coursework()
                    c.module = get_object_or_404(Module, code=line[0])
                    c.semester= str(line[2] or '')
                    c.week=str(line[3] or '')
                    c.weight=float(line[4] or 0.0)
                    c.information=str(line[5] or '')
                    c.save()
                elif line[1] == 'E':
                    e = Exam()
                    e.module = get_object_or_404(Module, code=line[0])
                    e.semester = str(line[2] or '')
                    e.week = str(line[3] or '')
                    e.weight = float(line[4] or 0.0)
                    e.information = str(line[5] or '')
                    e.duration = float(line[6] or 0.0)
                    e.save()
                elif line[1] == 'P':
                    p = Project()
                    p.module = get_object_or_404(Module, code=line[0])
                    p.semester= str(line[2] or '')
                    p.week=str(line[3] or '')
                    p.weight=float(line[4] or 0.0)
                    p.information=str(line[5] or '')
                    p.save()
    wb.close()
